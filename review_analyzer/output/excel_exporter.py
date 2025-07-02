# output/excel_exporter.py
"""엑셀 결과 출력 모듈"""

import pandas as pd
import os
from datetime import datetime
from typing import List, Tuple, Dict, Optional
from openpyxl.styles import Font, PatternFill, Alignment
from config.settings import ensure_output_dir, get_output_filename

class ExcelExporter:
    """엑셀 출력 클래스"""
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        엑셀 출력기 초기화
        
        Args:
            output_dir: 출력 디렉토리 경로
        """
        self.output_dir = ensure_output_dir(output_dir)
    
    def create_comprehensive_report(self, df: pd.DataFrame, 
                                  keywords_tfidf: List[Tuple[str, float]],
                                  keywords_krwordrank: Optional[List[Tuple[str, float]]],
                                  topics: Optional[List[List[str]]],
                                  sentiment_keywords: Dict[str, List[Tuple[str, float]]],
                                  initial_count: int, final_count: int,
                                  text_col: str, rating_col: Optional[str]) -> str:
        """
        종합 분석 리포트를 하나의 엑셀 시트로 생성
        
        Args:
            df: 분석된 DataFrame
            keywords_tfidf: TF-IDF 키워드 리스트
            keywords_krwordrank: KR-WordRank 키워드 리스트
            topics: 토픽 모델링 결과
            sentiment_keywords: 감성별 키워드
            initial_count: 처리 전 리뷰 수
            final_count: 처리 후 리뷰 수
            text_col: 텍스트 컬럼명
            rating_col: 평점 컬럼명
            
        Returns:
            출력 파일 경로
        """
        output_file = os.path.join(self.output_dir, get_output_filename())
        
        # 전체 분석 결과를 하나의 데이터프레임으로 구성
        all_data = self._build_report_data(
            df, keywords_tfidf, keywords_krwordrank, topics, sentiment_keywords,
            initial_count, final_count, text_col, rating_col
        )
        
        # 데이터프레임 생성 및 엑셀 저장
        df_report = pd.DataFrame(all_data, columns=['항목', '값', '설명/추가정보'])
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_report.to_excel(writer, sheet_name='쿠팡_리뷰_분석_리포트', index=False)
            self._style_worksheet(writer.sheets['쿠팡_리뷰_분석_리포트'], all_data)
        
        print(f"✅ 분석 결과가 하나의 엑셀 시트로 저장되었습니다:")
        print(f"📁 파일 경로: {output_file}")
        
        return output_file
    
    def create_detailed_report(self, df: pd.DataFrame,
                             keywords_tfidf: List[Tuple[str, float]],
                             keywords_krwordrank: Optional[List[Tuple[str, float]]],
                             topics: Optional[List[List[str]]],
                             sentiment_keywords: Dict[str, List[Tuple[str, float]]],
                             initial_count: int, final_count: int,
                             text_col: str, rating_col: Optional[str]) -> str:
        """
        상세 분석 리포트를 여러 시트로 생성
        
        Returns:
            출력 파일 경로
        """
        output_file = os.path.join(self.output_dir, get_output_filename("detailed_report"))
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. 분석 요약
            self._create_summary_sheet(writer, df, initial_count, final_count, text_col, rating_col)
            
            # 2. 감성 분석 결과
            self._create_sentiment_analysis_sheet(writer, df)
            
            # 3. 키워드 분석
            self._create_keyword_analysis_sheet(writer, keywords_tfidf, keywords_krwordrank)
            
            # 4. 토픽 모델링 (있는 경우)
            if topics:
                self._create_topic_modeling_sheet(writer, topics)
            
            # 5. 감성별 키워드
            self._create_sentiment_keywords_sheet(writer, sentiment_keywords)
            
            # 6. 상세 리뷰 데이터
            self._create_detailed_reviews_sheet(writer, df, text_col, rating_col)
            
            # 7. 통계 요약
            self._create_statistics_sheet(writer, df, rating_col)
        
        print(f"✅ 상세 분석 결과가 여러 시트로 저장되었습니다:")
        print(f"📁 파일 경로: {output_file}")
        
        return output_file
    
    def _build_report_data(self, df: pd.DataFrame, keywords_tfidf: List[Tuple[str, float]],
                          keywords_krwordrank: Optional[List[Tuple[str, float]]],
                          topics: Optional[List[List[str]]],
                          sentiment_keywords: Dict[str, List[Tuple[str, float]]],
                          initial_count: int, final_count: int,
                          text_col: str, rating_col: Optional[str]) -> List[List]:
        """리포트 데이터 구성"""
        all_data = []
        
        # 분석 개요
        all_data.extend([
            ['=== 쿠팡 리뷰 분석 리포트 ===', '', ''],
            ['분석 일시', datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ''],
            ['', '', '']
        ])
        
        # 기본 통계
        all_data.extend(self._get_basic_statistics(df, initial_count, final_count, text_col, rating_col))
        
        # 감성 분석 결과
        all_data.extend(self._get_sentiment_analysis_data(df))
        
        # TF-IDF 키워드 분석
        all_data.extend(self._get_tfidf_keywords_data(keywords_tfidf))
        
        # KR-WordRank 키워드 분석
        if keywords_krwordrank:
            all_data.extend(self._get_krwordrank_keywords_data(keywords_krwordrank))
        
        # 토픽 모델링 결과
        if topics:
            all_data.extend(self._get_topics_data(topics))
        
        # 감성별 키워드 분석
        all_data.extend(self._get_sentiment_keywords_data(sentiment_keywords))
        
        # 상세 통계
        all_data.extend(self._get_detailed_statistics(df))
        
        # 샘플 리뷰 데이터
        all_data.extend(self._get_sample_reviews_data(df, text_col))
        
        return all_data
    
    def _get_basic_statistics(self, df: pd.DataFrame, initial_count: int, final_count: int,
                            text_col: str, rating_col: Optional[str]) -> List[List]:
        """기본 통계 데이터"""
        data = [
            ['[기본 통계]', '', ''],
            ['총 리뷰 수 (원본)', f"{initial_count:,}개", '처리 전 전체 리뷰 개수'],
            ['분석된 리뷰 수', f"{final_count:,}개", '전처리 후 실제 분석된 리뷰 개수'],
            ['제거된 리뷰 수', f"{initial_count - final_count:,}개", '빈 내용 등으로 제거된 리뷰'],
            ['리뷰 텍스트 컬럼', text_col, '분석에 사용된 리뷰 텍스트 컬럼명'],
            ['평점 컬럼', rating_col if rating_col else '없음', '분석에 사용된 평점 컬럼명']
        ]
        
        # 텍스트 통계
        avg_length = df['cleaned_review'].str.len().mean()
        max_length = df['cleaned_review'].str.len().max()
        min_length = df['cleaned_review'].str.len().min()
        
        data.extend([
            ['평균 리뷰 길이', f"{avg_length:.1f}자", '전처리된 리뷰의 평균 글자 수'],
            ['최대 리뷰 길이', f"{max_length}자", '가장 긴 리뷰의 글자 수'],
            ['최소 리뷰 길이', f"{min_length}자", '가장 짧은 리뷰의 글자 수']
        ])
        
        # 평점 통계
        if rating_col and rating_col in df.columns:
            avg_rating = df[rating_col].mean()
            data.append(['평균 평점', f"{avg_rating:.2f}", '전체 리뷰의 평균 평점'])
            
            # 평점별 분포
            rating_counts = df[rating_col].value_counts().sort_index()
            for rating, count in rating_counts.items():
                ratio = count / len(df) * 100
                data.append([f"{rating}점 리뷰 수", f"{count}개 ({ratio:.1f}%)", '평점별 리뷰 분포'])
        
        data.append(['', '', ''])
        return data
    
    def _get_sentiment_analysis_data(self, df: pd.DataFrame) -> List[List]:
        """감성 분석 데이터"""
        data = [['[감성 분석 결과]', '', '']]
        
        if 'sentiment_rule' in df.columns:
            sentiment_counts = df['sentiment_rule'].value_counts()
            for sentiment in ['positive', 'negative', 'neutral']:
                count = sentiment_counts.get(sentiment, 0)
                ratio = count / len(df) * 100
                sentiment_kr = {'positive': '긍정', 'negative': '부정', 'neutral': '중립'}[sentiment]
                data.append([f"{sentiment_kr} 리뷰", f"{count:,}개 ({ratio:.1f}%)", '규칙 기반 감성분석 결과'])
        
        # 평점 기반 감성 분석 결과
        if 'sentiment_rating' in df.columns:
            data.extend([['', '', ''], ['[평점 기반 감성 분석]', '', '']])
            sentiment_rating_counts = df['sentiment_rating'].value_counts()
            for sentiment in ['positive', 'negative', 'neutral']:
                count = sentiment_rating_counts.get(sentiment, 0)
                ratio = count / len(df) * 100
                sentiment_kr = {'positive': '긍정', 'negative': '부정', 'neutral': '중립'}[sentiment]
                data.append([f"{sentiment_kr} 리뷰", f"{count:,}개 ({ratio:.1f}%)", '평점 기반 감성분석 결과'])
        
        data.append(['', '', ''])
        return data
    
    def _get_tfidf_keywords_data(self, keywords_tfidf: List[Tuple[str, float]]) -> List[List]:
        """TF-IDF 키워드 데이터"""
        data = [
            ['[TF-IDF 키워드 분석 TOP 20]', '', ''],
            ['순위', '키워드', 'TF-IDF 점수']
        ]
        
        for i, (word, score) in enumerate(keywords_tfidf[:20], 1):
            data.append([i, word, round(score, 4)])
        
        data.append(['', '', ''])
        return data
    
    def _get_krwordrank_keywords_data(self, keywords_krwordrank: List[Tuple[str, float]]) -> List[List]:
        """KR-WordRank 키워드 데이터"""
        data = [
            ['[KR-WordRank 키워드 분석 TOP 20]', '', ''],
            ['순위', '키워드', 'WordRank 점수']
        ]
        
        for i, (word, score) in enumerate(keywords_krwordrank[:20], 1):
            data.append([i, word, round(score, 1)])
        
        data.append(['', '', ''])
        return data
    
    def _get_topics_data(self, topics: List[List[str]]) -> List[List]:
        """토픽 모델링 데이터"""
        data = [
            ['[토픽 모델링 결과]', '', ''],
            ['토픽 번호', '주요 키워드', '설명']
        ]
        
        for i, topic_words in enumerate(topics, 1):
            keywords = ', '.join(topic_words[:8])
            data.append([f"토픽 {i}", keywords, f"토픽 {i}의 주요 키워드들"])
        
        data.append(['', '', ''])
        return data
    
    def _get_sentiment_keywords_data(self, sentiment_keywords: Dict[str, List[Tuple[str, float]]]) -> List[List]:
        """감성별 키워드 데이터"""
        data = [['[감성별 주요 키워드]', '', '']]
        
        for sentiment in ['positive', 'negative', 'neutral']:
            if sentiment in sentiment_keywords and sentiment_keywords[sentiment]:
                sentiment_kr = {'positive': '긍정', 'negative': '부정', 'neutral': '중립'}[sentiment]
                keywords_str = ', '.join([word for word, score in sentiment_keywords[sentiment][:8]])
                data.append([f"{sentiment_kr} 키워드", keywords_str, f"{sentiment_kr} 리뷰의 특징적 키워드"])
        
        data.append(['', '', ''])
        return data
    
    def _get_detailed_statistics(self, df: pd.DataFrame) -> List[List]:
        """상세 통계 데이터"""
        data = [['[상세 통계]', '', '']]
        
        # 텍스트 길이 통계
        length_stats = df['cleaned_review'].str.len().describe()
        data.extend([
            ['텍스트 길이 평균', f"{length_stats['mean']:.1f}자", ''],
            ['텍스트 길이 표준편차', f"{length_stats['std']:.1f}자", ''],
            ['텍스트 길이 최솟값', f"{length_stats['min']:.0f}자", ''],
            ['텍스트 길이 중간값', f"{length_stats['50%']:.0f}자", ''],
            ['텍스트 길이 최댓값', f"{length_stats['max']:.0f}자", '']
        ])
        
        # 키워드 수 통계
        if 'tokens' in df.columns:
            token_counts = df['tokens'].apply(len)
            token_stats = token_counts.describe()
            data.extend([
                ['키워드 수 평균', f"{token_stats['mean']:.1f}개", ''],
                ['키워드 수 중간값', f"{token_stats['50%']:.0f}개", ''],
                ['키워드 수 최댓값', f"{token_stats['max']:.0f}개", '']
            ])
        
        data.append(['', '', ''])
        return data
    
    def _get_sample_reviews_data(self, df: pd.DataFrame, text_col: str) -> List[List]:
        """샘플 리뷰 데이터"""
        data = [
            ['[샘플 리뷰 데이터 (상위 10개)]', '', ''],
            ['원본 리뷰', '규칙기반 감성', '추출된 키워드']
        ]
        
        sample_df = df.head(10)
        for idx, row in sample_df.iterrows():
            original_review = str(row[text_col])[:100] + "..." if len(str(row[text_col])) > 100 else str(row[text_col])
            sentiment = row.get('sentiment_rule', '없음')
            keywords = row.get('tokens_str', '')[:50] + "..." if len(str(row.get('tokens_str', ''))) > 50 else str(row.get('tokens_str', ''))
            data.append([original_review, sentiment, keywords])
        
        return data
    
    def _style_worksheet(self, worksheet, all_data: List[List]):
        """워크시트 스타일링"""
        # 컬럼 폭 조정
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 60
        
        # 헤더 스타일링
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, 4):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 섹션 헤더 스타일링
        section_font = Font(bold=True, size=11, color="0066CC")
        section_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for row in range(2, len(all_data) + 2):
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value and str(cell_value).startswith('[') and str(cell_value).endswith(']'):
                for col in range(1, 4):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = section_font
                    cell.fill = section_fill
    
    def _create_summary_sheet(self, writer, df: pd.DataFrame, initial_count: int, 
                            final_count: int, text_col: str, rating_col: Optional[str]):
        """요약 정보 시트 생성"""
        summary_data = [
            ['분석 항목', '값', '설명'],
            ['분석 일시', datetime.now().strftime("%Y-%m-%d %H:%M:%S"), '분석을 수행한 날짜와 시간'],
            ['총 리뷰 수 (원본)', f"{initial_count:,}개", '처리 전 전체 리뷰 개수'],
            ['분석된 리뷰 수', f"{final_count:,}개", '전처리 후 실제 분석된 리뷰 개수'],
            ['제거된 리뷰 수', f"{initial_count - final_count:,}개", '빈 내용 등으로 제거된 리뷰'],
            ['리뷰 텍스트 컬럼', text_col, '분석에 사용된 리뷰 텍스트 컬럼명'],
            ['평점 컬럼', rating_col if rating_col else '없음', '분석에 사용된 평점 컬럼명']
        ]
        
        # 기본 통계 추가
        avg_length = df['cleaned_review'].str.len().mean()
        summary_data.extend([
            ['평균 리뷰 길이', f"{avg_length:.1f}자", '전처리된 리뷰의 평균 글자 수']
        ])
        
        if rating_col and rating_col in df.columns:
            avg_rating = df[rating_col].mean()
            summary_data.append(['평균 평점', f"{avg_rating:.2f}", '전체 리뷰의 평균 평점'])
        
        # 감성 분포
        if 'sentiment_rule' in df.columns:
            sentiment_counts = df['sentiment_rule'].value_counts()
            summary_data.extend([
                ['긍정 리뷰 수', f"{sentiment_counts.get('positive', 0):,}개", '규칙 기반 감성분석 결과'],
                ['부정 리뷰 수', f"{sentiment_counts.get('negative', 0):,}개", '규칙 기반 감성분석 결과'],
                ['중립 리뷰 수', f"{sentiment_counts.get('neutral', 0):,}개", '규칙 기반 감성분석 결과']
            ])
        
        summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
        summary_df.to_excel(writer, sheet_name='1_분석요약', index=False)
    
    def _create_sentiment_analysis_sheet(self, writer, df: pd.DataFrame):
        """감성 분석 결과 시트 생성"""
        sentiment_data = [['감성', '개수', '비율(%)', '분석방법']]
        
        # 규칙 기반 감성 분석
        if 'sentiment_rule' in df.columns:
            sentiment_rule_counts = df['sentiment_rule'].value_counts()
            sentiment_rule_ratio = (sentiment_rule_counts / len(df) * 100).round(1)
            
            for sentiment in ['positive', 'negative', 'neutral']:
                count = sentiment_rule_counts.get(sentiment, 0)
                ratio = sentiment_rule_ratio.get(sentiment, 0)
                sentiment_data.append([sentiment, count, ratio, '규칙기반'])
        
        # 평점 기반 감성 분석
        if 'sentiment_rating' in df.columns:
            sentiment_rating_counts = df['sentiment_rating'].value_counts()
            sentiment_rating_ratio = (sentiment_rating_counts / len(df) * 100).round(1)
            
            for sentiment in ['positive', 'negative', 'neutral']:
                count = sentiment_rating_counts.get(sentiment, 0)
                ratio = sentiment_rating_ratio.get(sentiment, 0)
                sentiment_data.append([sentiment, count, ratio, '평점기반'])
        
        sentiment_df = pd.DataFrame(sentiment_data[1:], columns=sentiment_data[0])
        sentiment_df.to_excel(writer, sheet_name='2_감성분석결과', index=False)
    
    def _create_keyword_analysis_sheet(self, writer, keywords_tfidf: List[Tuple[str, float]],
                                     keywords_krwordrank: Optional[List[Tuple[str, float]]]):
        """키워드 분석 시트 생성"""
        # TF-IDF 키워드
        tfidf_data = [['순위', '키워드', 'TF-IDF 점수']]
        for i, (word, score) in enumerate(keywords_tfidf[:30], 1):
            tfidf_data.append([i, word, round(score, 4)])
        
        tfidf_df = pd.DataFrame(tfidf_data[1:], columns=tfidf_data[0])
        
        # KR-WordRank 키워드
        if keywords_krwordrank:
            krwordrank_data = [['순위', '키워드', 'WordRank 점수']]
            for i, (word, score) in enumerate(keywords_krwordrank[:30], 1):
                krwordrank_data.append([i, word, round(score, 1)])
            
            krwordrank_df = pd.DataFrame(krwordrank_data[1:], columns=krwordrank_data[0])
            
            # 두 결과를 나란히 배치
            combined_df = pd.concat([tfidf_df, pd.DataFrame([''] * len(tfidf_df)), krwordrank_df], axis=1)
            combined_df.columns = ['TF-IDF 순위', 'TF-IDF 키워드', 'TF-IDF 점수', '', 
                                  'WordRank 순위', 'WordRank 키워드', 'WordRank 점수']
        else:
            combined_df = tfidf_df
        
        combined_df.to_excel(writer, sheet_name='3_키워드분석', index=False)
    
    def _create_topic_modeling_sheet(self, writer, topics: List[List[str]]):
        """토픽 모델링 시트 생성"""
        topic_data = [['토픽 번호', '주요 키워드', '키워드 설명']]
        
        for i, topic_words in enumerate(topics, 1):
            keywords = ', '.join(topic_words[:10])
            description = f"토픽 {i}의 주요 키워드들"
            topic_data.append([f"토픽 {i}", keywords, description])
        
        topic_df = pd.DataFrame(topic_data[1:], columns=topic_data[0])
        topic_df.to_excel(writer, sheet_name='4_토픽모델링', index=False)
    
    def _create_sentiment_keywords_sheet(self, writer, sentiment_keywords: Dict[str, List[Tuple[str, float]]]):
        """감성별 키워드 시트 생성"""
        sentiment_keywords_data = [['감성', '순위', '키워드', 'TF-IDF 점수']]
        
        for sentiment in ['positive', 'negative', 'neutral']:
            if sentiment in sentiment_keywords and sentiment_keywords[sentiment]:
                for i, (word, score) in enumerate(sentiment_keywords[sentiment], 1):
                    sentiment_keywords_data.append([sentiment, i, word, round(score, 4)])
        
        sentiment_keywords_df = pd.DataFrame(sentiment_keywords_data[1:], columns=sentiment_keywords_data[0])
        sentiment_keywords_df.to_excel(writer, sheet_name='5_감성별키워드', index=False)
    
    def _create_detailed_reviews_sheet(self, writer, df: pd.DataFrame, text_col: str, rating_col: Optional[str]):
        """상세 리뷰 데이터 시트 생성"""
        columns_to_include = [text_col, 'cleaned_review', 'sentiment_rule']
        
        if rating_col and rating_col in df.columns:
            columns_to_include.extend([rating_col, 'sentiment_rating'])
        
        if 'tokens_str' in df.columns:
            columns_to_include.append('tokens_str')
        
        detailed_df = df[columns_to_include].copy()
        
        # 컬럼명 변경
        column_names = {
            text_col: '원본_리뷰',
            'cleaned_review': '전처리된_리뷰',
            'sentiment_rule': '규칙기반_감성',
            'tokens_str': '추출된_키워드'
        }
        
        if rating_col and rating_col in df.columns:
            column_names[rating_col] = '평점'
            column_names['sentiment_rating'] = '평점기반_감성'
        
        detailed_df = detailed_df.rename(columns=column_names)
        
        # 엑셀 시트 크기 제한을 고려하여 최대 10000개까지만
        if len(detailed_df) > 10000:
            detailed_df = detailed_df.head(10000)
        
        detailed_df.to_excel(writer, sheet_name='6_상세리뷰데이터', index=False)
    
    def _create_statistics_sheet(self, writer, df: pd.DataFrame, rating_col: Optional[str]):
        """통계 요약 시트 생성"""
        stats_data = [['구분', '항목', '값']]
        
        # 텍스트 길이 통계
        length_stats = df['cleaned_review'].str.len().describe()
        stats_data.extend([
            ['텍스트 길이', '평균', f"{length_stats['mean']:.1f}자"],
            ['텍스트 길이', '표준편차', f"{length_stats['std']:.1f}자"],
            ['텍스트 길이', '최솟값', f"{length_stats['min']:.0f}자"],
            ['텍스트 길이', '25% 분위수', f"{length_stats['25%']:.0f}자"],
            ['텍스트 길이', '중간값', f"{length_stats['50%']:.0f}자"],
            ['텍스트 길이', '75% 분위수', f"{length_stats['75%']:.0f}자"],
            ['텍스트 길이', '최댓값', f"{length_stats['max']:.0f}자"]
        ])
        
        # 평점 통계
        if rating_col and rating_col in df.columns:
            rating_stats = df[rating_col].describe()
            stats_data.extend([
                ['평점', '평균', f"{rating_stats['mean']:.2f}"],
                ['평점', '표준편차', f"{rating_stats['std']:.2f}"],
                ['평점', '최솟값', f"{rating_stats['min']:.0f}"],
                ['평점', '중간값', f"{rating_stats['50%']:.0f}"],
                ['평점', '최댓값', f"{rating_stats['max']:.0f}"]
            ])
        
        # 토큰 수 통계
        if 'tokens' in df.columns:
            token_counts = df['tokens'].apply(len)
            token_stats = token_counts.describe()
            stats_data.extend([
                ['키워드 수', '평균', f"{token_stats['mean']:.1f}개"],
                ['키워드 수', '중간값', f"{token_stats['50%']:.0f}개"],
                ['키워드 수', '최댓값', f"{token_stats['max']:.0f}개"]
            ])
        
        stats_df = pd.DataFrame(stats_data[1:], columns=stats_data[0])
        stats_df.to_excel(writer, sheet_name='7_통계요약', index=False)