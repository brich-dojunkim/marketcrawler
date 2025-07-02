# 완성된 한국어 상품 리뷰 분석 시스템
# 쿠팡 리뷰 데이터 분석 및 엑셀 결과 출력

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from collections import Counter
import re
import warnings
from datetime import datetime
import os
warnings.filterwarnings('ignore')

# 한국어 자연어처리 라이브러리
# KoNLPy 임포트 시 Java 오류 처리
try:
    from konlpy.tag import Okt, Kkma, Komoran
    KONLPY_AVAILABLE = True
    print("KoNLPy 사용 가능")
except Exception as e:
    print(f"KoNLPy 사용 불가: {e}")
    print("Java가 필요합니다. 설치 가이드를 확인하세요.")
    KONLPY_AVAILABLE = False

from kiwipiepy import Kiwi

# PyKoSpacing 임포트 시 오류 처리 (TensorFlow 의존성)
try:
    from pykospacing import Spacing
    PYKOSPACING_AVAILABLE = True
    print("PyKoSpacing 사용 가능")
except ImportError as e:
    print(f"PyKoSpacing 사용 불가: {e}")
    print("TensorFlow가 필요합니다. 'pip install tensorflow' 설치 후 재시도하세요.")
    PYKOSPACING_AVAILABLE = False

# 비지도 학습 기반
from soynlp.word import WordExtractor
from soynlp.tokenizer import LTokenizer
from krwordrank.word import KRWordRank

# 딥러닝 기반 (KoBERT) - 선택적 임포트
try:
    import torch
    from transformers import BertTokenizer, BertForSequenceClassification
    TRANSFORMERS_AVAILABLE = True
    print("PyTorch/Transformers 사용 가능")
except ImportError:
    print("PyTorch/Transformers를 설치하지 못했습니다. 규칙 기반 감성 분석을 사용합니다.")
    TRANSFORMERS_AVAILABLE = False

# 시각화 및 분석
from wordcloud import WordCloud
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from sklearn.decomposition import LatentDirichletAllocation

# =============================================================================
# ReviewAnalyzer 클래스
# =============================================================================

class ReviewAnalyzer:
    """상품 리뷰 분석 통합 클래스"""
    
    def __init__(self):
        # 전역 변수 참조
        global KONLPY_AVAILABLE
        
        # 형태소 분석기 초기화
        self.kiwi = Kiwi()  # Kiwi는 Java 불필요
        
        # KoNLPy 초기화 (Java 사용 가능한 경우)
        if KONLPY_AVAILABLE:
            try:
                self.okt = Okt()
                self.kkma = Kkma()
                print("KoNLPy 형태소 분석기 초기화 완료")
            except Exception as e:
                print(f"KoNLPy 초기화 실패: {e}")
                KONLPY_AVAILABLE = False
                self.okt = None
                self.kkma = None
        else:
            self.okt = None
            self.kkma = None
        
        # PyKoSpacing 초기화 (사용 가능한 경우)
        if PYKOSPACING_AVAILABLE:
            self.spacing = Spacing()
        else:
            self.spacing = None
        
        # 불용어 정의
        self.stopwords = [
            '은', '는', '이', '가', '을', '를', '의', '에', '에서', '로', '으로',
            '과', '와', '도', '만', '부터', '까지', '뿐', '아니라', '처럼', '같이',
            '그', '저', '이', '그것', '저것', '것', '수', '곳', '때', '점',
            '정말', '너무', '진짜', '완전', '매우', '아주', '좀', '조금'
        ]
    
    def load_data(self, file_path, text_column='review', rating_column='rating'):
        """
        리뷰 데이터 로딩
        
        Args:
            file_path: 데이터 파일 경로
            text_column: 리뷰 텍스트 컬럼명
            rating_column: 평점 컬럼명
        """
        self.df = pd.read_csv(file_path)
        self.text_col = text_column
        self.rating_col = rating_column
        
        print(f"데이터 로딩 완료: {len(self.df)}개 리뷰")
        print(f"컬럼: {list(self.df.columns)}")
        
        return self.df

    def preprocess_text(self, text):
        """
        텍스트 전처리
        - 띄어쓰기 교정 (PyKoSpacing 사용 가능시)
        - 특수문자 제거
        - 정규화
        """
        if pd.isna(text):
            return ""
        
        # 띄어쓰기 교정 (PyKoSpacing이 사용 가능한 경우만)
        if self.spacing is not None:
            try:
                text = self.spacing(text)
            except Exception as e:
                print(f"띄어쓰기 교정 중 오류: {e}")
                pass  # 오류 시 원본 텍스트 사용
        else:
            # 간단한 띄어쓰기 규칙 적용
            text = self.simple_spacing_correction(text)
        
        # 특수문자 제거 (한글, 영문, 숫자, 공백만 유지)
        text = re.sub(r'[^가-힣a-zA-Z0-9\s]', '', text)
        
        # 연속된 공백 제거
        text = re.sub(r'\s+', ' ', text)
        
        # 앞뒤 공백 제거
        text = text.strip()
        
        return text
    
    def simple_spacing_correction(self, text):
        """
        간단한 띄어쓰기 규칙 적용 (PyKoSpacing 대체)
        """
        # 기본적인 띄어쓰기 규칙들
        spacing_rules = [
            # 조사 앞에 띄어쓰기
            (r'([가-힣])([은는이가을를])([가-힣])', r'\1 \2 \3'),
            # 어미 앞에 띄어쓰기  
            (r'([가-힣])([습니다해요])([가-힣])', r'\1\2 \3'),
            # 숫자와 단위 사이
            (r'([0-9])([가-힣])', r'\1 \2'),
            # 영어와 한글 사이
            (r'([a-zA-Z])([가-힣])', r'\1 \2'),
            (r'([가-힣])([a-zA-Z])', r'\1 \2'),
        ]
        
        for pattern, replacement in spacing_rules:
            text = re.sub(pattern, replacement, text)
        
        return text
    
    def clean_reviews(self):
        """전체 리뷰 데이터 전처리"""
        print("리뷰 데이터 전처리 중...")
        
        self.df['cleaned_review'] = self.df[self.text_col].apply(self.preprocess_text)
        
        # 빈 리뷰 제거
        self.df = self.df[self.df['cleaned_review'].str.len() > 0].reset_index(drop=True)
        
        print(f"전처리 완료: {len(self.df)}개 리뷰")
        
        return self.df

    def tokenize_with_okt(self, text, pos_filter=['Noun', 'Verb', 'Adjective']):
        """OKT를 이용한 형태소 분석 (Java 필요)"""
        if not text:
            return []
        
        if not KONLPY_AVAILABLE or self.okt is None:
            print("KoNLPy를 사용할 수 없습니다. Kiwi를 사용하세요.")
            return self.tokenize_with_kiwi(text)
        
        try:
            tokens = self.okt.pos(text, stem=True)
            filtered_tokens = [
                word for word, pos in tokens 
                if pos in pos_filter and word not in self.stopwords and len(word) > 1
            ]
            return filtered_tokens
        except Exception as e:
            print(f"OKT 분석 오류: {e}")
            return self.tokenize_with_kiwi(text)
    
    def tokenize_with_kiwi(self, text, pos_filter=['NNG', 'NNP', 'VA', 'VV']):
        """Kiwi를 이용한 형태소 분석"""
        if not text:
            return []
        
        tokens = self.kiwi.tokenize(text)
        filtered_tokens = [
            token.form for token in tokens
            if token.tag in pos_filter and token.form not in self.stopwords and len(token.form) > 1
        ]
        
        return filtered_tokens
    
    def extract_words_with_soynlp(self, min_count=10, max_length=10):
        """soynlp를 이용한 단어 추출"""
        print("soynlp로 단어 추출 중...")
        
        texts = self.df['cleaned_review'].tolist()
        
        word_extractor = WordExtractor(
            min_count=min_count,
            min_length=2,
            max_length=max_length
        )
        
        word_extractor.train(texts)
        words = word_extractor.extract()
        
        # 단어 점수가 높은 순으로 정렬
        word_scores = {
            word: score.cohesion_forward * score.right_branching_entropy
            for word, score in words.items()
        }
        
        return word_scores
    
    def tokenize_reviews(self, method='kiwi'):
        """전체 리뷰 토큰화"""
        print(f"{method}로 리뷰 토큰화 중...")
        
        if method == 'okt':
            self.df['tokens'] = self.df['cleaned_review'].apply(self.tokenize_with_okt)
        elif method == 'kiwi':
            self.df['tokens'] = self.df['cleaned_review'].apply(self.tokenize_with_kiwi)
        
        # 토큰화된 결과를 문자열로 변환 (일부 분석에서 필요)
        self.df['tokens_str'] = self.df['tokens'].apply(lambda x: ' '.join(x))
        
        print("토큰화 완료")
        return self.df

    def load_sentiment_dict(self, dict_path=None):
        """한국어 감성사전 로딩"""
        # 확장된 감성사전
        positive_words = [
            '좋다', '훌륭하다', '만족', '추천', '최고', '완벽', '우수', '탁월',
            '뛰어나다', '멋지다', '성공', '효과', '품질', '가성비', '깔끔하다',
            '예쁘다', '괜찮다', '마음에들다', '든든하다', '편리하다', '유용하다',
            '빠르다', '정확하다', '친절하다', '싸다', '저렴하다', '합리적',
            '효율적', '신속하다', '만족스럽다', '감사하다', '고맙다', '도움',
            '편하다', '쉽다', '간편하다', '실용적', '경제적', '알뜰하다'
        ]
        
        negative_words = [
            '나쁘다', '별로', '최악', '실망', '불만', '문제', '부족', '아쉽다',
            '후회', '비추', '개선', '단점', '불편', '오류', '느리다', '비싸다',
            '어렵다', '복잡하다', '귀찮다', '짜증', '화나다', '속상하다',
            '걱정', '불안', '고장', '파손', '결함', '흠집', '더럽다',
            '냄새', '시끄럽다', '무겁다', '작다', '크다', '딱딱하다', '질기다'
        ]
        
        self.sentiment_dict = {
            'positive': positive_words,
            'negative': negative_words
        }
        
        return self.sentiment_dict
    
    def analyze_sentiment_rule_based(self, tokens):
        """개선된 규칙 기반 감성 분석"""
        if not hasattr(self, 'sentiment_dict'):
            self.load_sentiment_dict()
        
        pos_score = sum(1 for token in tokens if token in self.sentiment_dict['positive'])
        neg_score = sum(1 for token in tokens if token in self.sentiment_dict['negative'])
        
        # 점수 차이로 중립 판단
        if abs(pos_score - neg_score) <= 1 and (pos_score + neg_score) <= 2:
            return 'neutral'
        elif pos_score > neg_score:
            return 'positive'
        elif neg_score > pos_score:
            return 'negative'
        else:
            return 'neutral'
    
    def analyze_sentiment_rating_based(self, rating, threshold=3):
        """평점 기반 감성 라벨링"""
        if rating > threshold:
            return 'positive'
        elif rating < threshold:
            return 'negative'
        else:
            return 'neutral'
    
    def create_sentiment_labels(self, method='rating'):
        """감성 라벨 생성"""
        print(f"{method} 방식으로 감성 라벨 생성 중...")
        
        if method == 'rating':
            self.df['sentiment'] = self.df[self.rating_col].apply(self.analyze_sentiment_rating_based)
        elif method == 'rule':
            self.df['sentiment'] = self.df['tokens'].apply(self.analyze_sentiment_rule_based)
        
        print("감성 라벨 생성 완료")
        print(self.df['sentiment'].value_counts())
        
        return self.df

    def extract_keywords_tfidf(self, max_features=100, ngram_range=(1, 2)):
        """TF-IDF 기반 키워드 추출"""
        print("TF-IDF로 키워드 추출 중...")
        
        vectorizer = TfidfVectorizer(
            max_features=max_features,
            ngram_range=ngram_range,
            token_pattern=r'\b\w+\b'
        )
        
        tfidf_matrix = vectorizer.fit_transform(self.df['tokens_str'])
        feature_names = vectorizer.get_feature_names_out()
        
        # 단어별 TF-IDF 점수 계산
        tfidf_scores = tfidf_matrix.sum(axis=0).A1
        keyword_scores = list(zip(feature_names, tfidf_scores))
        keyword_scores.sort(key=lambda x: x[1], reverse=True)
        
        self.tfidf_keywords = keyword_scores
        
        return keyword_scores
    
    def extract_keywords_krwordrank(self, min_count=5, max_length=10):
        """KR-WordRank로 키워드 추출"""
        print("KR-WordRank로 키워드 추출 중...")
        
        texts = self.df['cleaned_review'].tolist()
        
        wordrank_extractor = KRWordRank(
            min_count=min_count,
            max_length=max_length
        )
        
        beta = 0.85
        max_iter = 10
        
        keywords, rank, graph = wordrank_extractor.extract(texts, beta, max_iter)
        
        # 키워드를 점수순으로 정렬
        sorted_keywords = sorted(keywords.items(), key=lambda x: x[1], reverse=True)
        
        self.krwordrank_keywords = sorted_keywords
        
        return sorted_keywords

    def topic_modeling_lda(self, n_topics=5, max_features=100):
        """LDA 토픽 모델링"""
        print(f"LDA로 {n_topics}개 토픽 추출 중...")
        
        vectorizer = TfidfVectorizer(
            max_features=max_features,
            token_pattern=r'\b\w+\b'
        )
        
        tfidf_matrix = vectorizer.fit_transform(self.df['tokens_str'])
        
        lda = LatentDirichletAllocation(
            n_components=n_topics,
            random_state=42
        )
        
        lda.fit(tfidf_matrix)
        
        # 토픽별 주요 단어 추출
        feature_names = vectorizer.get_feature_names_out()
        topics = []
        
        for topic_idx, topic in enumerate(lda.components_):
            top_words_idx = topic.argsort()[-10:][::-1]
            top_words = [feature_names[i] for i in top_words_idx]
            topics.append(top_words)
        
        self.topics = topics
        
        return topics

    def plot_sentiment_distribution(self):
        """감성 분포 시각화"""
        plt.figure(figsize=(10, 6))
        
        # 감성별 분포
        plt.subplot(1, 2, 1)
        sentiment_counts = self.df['sentiment'].value_counts()
        plt.pie(sentiment_counts.values, labels=sentiment_counts.index, autopct='%1.1f%%')
        plt.title('감성 분포')
        
        # 평점별 분포
        plt.subplot(1, 2, 2)
        plt.hist(self.df[self.rating_col], bins=5, edgecolor='black')
        plt.xlabel('평점')
        plt.ylabel('리뷰 수')
        plt.title('평점 분포')
        
        plt.tight_layout()
        plt.show()
    
    def create_wordcloud(self, sentiment=None, max_words=100):
        """워드클라우드 생성"""
        if sentiment:
            text_data = self.df[self.df['sentiment'] == sentiment]['tokens_str']
            title = f'{sentiment.title()} 리뷰 워드클라우드'
        else:
            text_data = self.df['tokens_str']
            title = '전체 리뷰 워드클라우드'
        
        text = ' '.join(text_data)
        
        # macOS에서 한글 폰트 경로 설정
        font_paths = [
            '/System/Library/Fonts/NanumGothic.ttc',  # macOS 기본 경로
            '/Library/Fonts/NanumGothic.ttf',
            '/System/Library/Fonts/Arial Unicode MS.ttf',  # 대안 폰트
            '/System/Library/Fonts/AppleGothic.ttf'
        ]
        
        font_path = None
        for path in font_paths:
            try:
                import os
                if os.path.exists(path):
                    font_path = path
                    break
            except:
                continue
        
        wordcloud_params = {
            'width': 800,
            'height': 400,
            'background_color': 'white',
            'max_words': max_words
        }
        
        if font_path:
            wordcloud_params['font_path'] = font_path
        
        wordcloud = WordCloud(**wordcloud_params).generate(text)
        
        plt.figure(figsize=(12, 6))
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.axis('off')
        plt.title(title, fontsize=16)
        plt.show()
    
    def plot_keyword_ranking(self, method='tfidf', top_n=20):
        """키워드 랭킹 시각화"""
        if method == 'tfidf' and hasattr(self, 'tfidf_keywords'):
            keywords = self.tfidf_keywords[:top_n]
            title = 'TF-IDF 키워드 랭킹'
        elif method == 'krwordrank' and hasattr(self, 'krwordrank_keywords'):
            keywords = self.krwordrank_keywords[:top_n]
            title = 'KR-WordRank 키워드 랭킹'
        else:
            print(f"{method} 키워드가 없습니다. 먼저 키워드 추출을 실행하세요.")
            return
        
        words = [item[0] for item in keywords]
        scores = [item[1] for item in keywords]
        
        plt.figure(figsize=(12, 8))
        plt.barh(range(len(words)), scores)
        plt.yticks(range(len(words)), words)
        plt.xlabel('점수')
        plt.title(title)
        plt.gca().invert_yaxis()
        plt.tight_layout()
        plt.show()

    def generate_analysis_report(self):
        """분석 결과 종합 리포트"""
        print("="*50)
        print("상품 리뷰 분석 리포트")
        print("="*50)
        
        # 기본 통계
        print(f"\n1. 기본 통계")
        print(f"   - 총 리뷰 수: {len(self.df):,}개")
        print(f"   - 평균 평점: {self.df[self.rating_col].mean():.2f}")
        print(f"   - 평균 리뷰 길이: {self.df['cleaned_review'].str.len().mean():.1f}자")
        
        # 감성 분석 결과
        if 'sentiment' in self.df.columns:
            print(f"\n2. 감성 분석 결과")
            sentiment_counts = self.df['sentiment'].value_counts()
            for sentiment, count in sentiment_counts.items():
                ratio = count / len(self.df) * 100
                print(f"   - {sentiment}: {count:,}개 ({ratio:.1f}%)")
        
        # 주요 키워드
        if hasattr(self, 'tfidf_keywords'):
            print(f"\n3. 주요 키워드 (TF-IDF)")
            for i, (word, score) in enumerate(self.tfidf_keywords[:10], 1):
                print(f"   {i:2d}. {word} ({score:.3f})")
        
        # 토픽 모델링 결과
        if hasattr(self, 'topics'):
            print(f"\n4. 주요 토픽")
            for i, topic_words in enumerate(self.topics, 1):
                print(f"   토픽 {i}: {', '.join(topic_words[:5])}")
        
        print("="*50)

# =============================================================================
# 엑셀 결과 출력 함수들
# =============================================================================

def analyze_coupang_reviews():
    """쿠팡 리뷰 데이터 분석 실행"""
    
    # 1. 데이터 로딩 및 확인
    file_path = '/Users/brich/Desktop/marketcrawler/output/coupang_reviews_20250701_180647.csv'
    
    try:
        # CSV 파일 읽기
        df = pd.read_csv(file_path)
        return df
        
    except FileNotFoundError:
        print("❌ 파일을 찾을 수 없습니다. 경로를 확인해주세요.")
        return None
    except Exception as e:
        print(f"❌ 데이터 로딩 중 오류: {e}")
        return None

def identify_columns(df):
    """데이터 컬럼 분석 및 리뷰/평점 컬럼 식별"""
    
    text_columns = []
    rating_columns = []
    
    for col in df.columns:
        # 텍스트 컬럼 추정 (긴 문자열, 높은 다양성)
        if df[col].dtype == 'object':
            avg_length = df[col].dropna().astype(str).str.len().mean()
            if avg_length > 10:  # 평균 길이가 10자 이상
                text_columns.append(col)
        
        # 평점 컬럼 추정 (숫자형, 1-5 또는 1-10 범위)
        if df[col].dtype in ['int64', 'float64']:
            min_val = df[col].min()
            max_val = df[col].max()
            if 1 <= min_val and max_val <= 10:  # 1-10 범위의 숫자
                rating_columns.append(col)
    
    return text_columns, rating_columns

def run_analysis_with_identified_columns(df, text_col, rating_col=None):
    """식별된 컬럼으로 분석 실행 및 엑셀 결과 생성"""
    
    analyzer = ReviewAnalyzer()
    
    # 데이터 설정
    analyzer.df = df.copy()
    analyzer.text_col = text_col
    analyzer.rating_col = rating_col
    
    # 분석 파이프라인 실행
    # 1. 텍스트 전처리
    analyzer.df['cleaned_review'] = analyzer.df[text_col].apply(analyzer.preprocess_text)
    
    # 빈 리뷰 제거
    initial_count = len(analyzer.df)
    analyzer.df = analyzer.df[analyzer.df['cleaned_review'].str.len() > 0].reset_index(drop=True)
    final_count = len(analyzer.df)
    
    # 2. 형태소 분석
    analyzer.df['tokens'] = analyzer.df['cleaned_review'].apply(analyzer.tokenize_with_kiwi)
    analyzer.df['tokens_str'] = analyzer.df['tokens'].apply(lambda x: ' '.join(x))
    
    # 3. 감성 분석
    analyzer.df['sentiment_rule'] = analyzer.df['tokens'].apply(analyzer.analyze_sentiment_rule_based)
    
    if rating_col and rating_col in df.columns:
        analyzer.df['sentiment_rating'] = analyzer.df[rating_col].apply(analyzer.analyze_sentiment_rating_based)
    
    # 4. 키워드 추출
    keywords_tfidf = analyzer.extract_keywords_tfidf(max_features=50)
    
    try:
        keywords_krwordrank = analyzer.extract_keywords_krwordrank(min_count=3)
    except Exception as e:
        keywords_krwordrank = None
    
    # 5. 토픽 모델링
    try:
        topics = analyzer.topic_modeling_lda(n_topics=5, max_features=50)
    except Exception as e:
        topics = None
    
    # 엑셀 결과 생성
    create_excel_report(analyzer, keywords_tfidf, keywords_krwordrank, topics, 
                       initial_count, final_count, text_col, rating_col)
    
    return analyzer

def create_excel_report(analyzer, keywords_tfidf, keywords_krwordrank, topics, 
                       initial_count, final_count, text_col, rating_col):
    """분석 결과를 하나의 엑셀 시트로 저장"""
    
    # 현재 시간으로 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = '/Users/brich/Desktop/marketcrawler/output'
    
    # 출력 디렉토리가 없으면 생성
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    output_file = f"{output_dir}/coupang_analysis_report_{timestamp}.xlsx"
    
    # 전체 분석 결과를 하나의 데이터프레임으로 구성
    all_data = []
    
    # 1. 분석 개요
    all_data.append(['=== 쿠팡 리뷰 분석 리포트 ===', '', ''])
    all_data.append(['분석 일시', datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ''])
    all_data.append(['', '', ''])
    
    # 2. 기본 통계
    all_data.append(['[기본 통계]', '', ''])
    all_data.append(['총 리뷰 수 (원본)', f"{initial_count:,}개", '처리 전 전체 리뷰 개수'])
    all_data.append(['분석된 리뷰 수', f"{final_count:,}개", '전처리 후 실제 분석된 리뷰 개수'])
    all_data.append(['제거된 리뷰 수', f"{initial_count - final_count:,}개", '빈 내용 등으로 제거된 리뷰'])
    all_data.append(['리뷰 텍스트 컬럼', text_col, '분석에 사용된 리뷰 텍스트 컬럼명'])
    all_data.append(['평점 컬럼', rating_col if rating_col else '없음', '분석에 사용된 평점 컬럼명'])
    
    # 텍스트 통계
    avg_length = analyzer.df['cleaned_review'].str.len().mean()
    max_length = analyzer.df['cleaned_review'].str.len().max()
    min_length = analyzer.df['cleaned_review'].str.len().min()
    
    all_data.append(['평균 리뷰 길이', f"{avg_length:.1f}자", '전처리된 리뷰의 평균 글자 수'])
    all_data.append(['최대 리뷰 길이', f"{max_length}자", '가장 긴 리뷰의 글자 수'])
    all_data.append(['최소 리뷰 길이', f"{min_length}자", '가장 짧은 리뷰의 글자 수'])
    
    # 평점 통계
    if rating_col and rating_col in analyzer.df.columns:
        avg_rating = analyzer.df[rating_col].mean()
        all_data.append(['평균 평점', f"{avg_rating:.2f}", '전체 리뷰의 평균 평점'])
        
        # 평점별 분포
        rating_counts = analyzer.df[rating_col].value_counts().sort_index()
        for rating, count in rating_counts.items():
            ratio = count / len(analyzer.df) * 100
            all_data.append([f"{rating}점 리뷰 수", f"{count}개 ({ratio:.1f}%)", '평점별 리뷰 분포'])
    
    all_data.append(['', '', ''])
    
    # 3. 감성 분석 결과
    all_data.append(['[감성 분석 결과]', '', ''])
    sentiment_counts = analyzer.df['sentiment_rule'].value_counts()
    
    for sentiment in ['positive', 'negative', 'neutral']:
        count = sentiment_counts.get(sentiment, 0)
        ratio = count / len(analyzer.df) * 100
        sentiment_kr = {'positive': '긍정', 'negative': '부정', 'neutral': '중립'}[sentiment]
        all_data.append([f"{sentiment_kr} 리뷰", f"{count:,}개 ({ratio:.1f}%)", '규칙 기반 감성분석 결과'])
    
    # 평점 기반 감성 분석 결과 (있는 경우)
    if 'sentiment_rating' in analyzer.df.columns:
        all_data.append(['', '', ''])
        all_data.append(['[평점 기반 감성 분석]', '', ''])
        sentiment_rating_counts = analyzer.df['sentiment_rating'].value_counts()
        
        for sentiment in ['positive', 'negative', 'neutral']:
            count = sentiment_rating_counts.get(sentiment, 0)
            ratio = count / len(analyzer.df) * 100
            sentiment_kr = {'positive': '긍정', 'negative': '부정', 'neutral': '중립'}[sentiment]
            all_data.append([f"{sentiment_kr} 리뷰", f"{count:,}개 ({ratio:.1f}%)", '평점 기반 감성분석 결과'])
    
    all_data.append(['', '', ''])
    
    # 4. TF-IDF 키워드 분석
    all_data.append(['[TF-IDF 키워드 분석 TOP 20]', '', ''])
    all_data.append(['순위', '키워드', 'TF-IDF 점수'])
    
    for i, (word, score) in enumerate(keywords_tfidf[:20], 1):
        all_data.append([i, word, round(score, 4)])
    
    all_data.append(['', '', ''])
    
    # 5. KR-WordRank 키워드 분석 (있는 경우)
    if keywords_krwordrank:
        all_data.append(['[KR-WordRank 키워드 분석 TOP 20]', '', ''])
        all_data.append(['순위', '키워드', 'WordRank 점수'])
        
        for i, (word, score) in enumerate(keywords_krwordrank[:20], 1):
            all_data.append([i, word, round(score, 1)])
        
        all_data.append(['', '', ''])
    
    # 6. 토픽 모델링 결과
    if topics:
        all_data.append(['[토픽 모델링 결과]', '', ''])
        all_data.append(['토픽 번호', '주요 키워드', '설명'])
        
        for i, topic_words in enumerate(topics, 1):
            keywords = ', '.join(topic_words[:8])
            all_data.append([f"토픽 {i}", keywords, f"토픽 {i}의 주요 키워드들"])
        
        all_data.append(['', '', ''])
    
    # 7. 감성별 키워드 분석
    all_data.append(['[감성별 주요 키워드]', '', ''])
    
    for sentiment in ['positive', 'negative', 'neutral']:
        if sentiment in analyzer.df['sentiment_rule'].values:
            sentiment_data = analyzer.df[analyzer.df['sentiment_rule'] == sentiment]
            if len(sentiment_data) > 0:
                sentiment_text = ' '.join(sentiment_data['tokens_str'])
                vectorizer = TfidfVectorizer(max_features=10, token_pattern=r'\b\w+\b')
                try:
                    tfidf_matrix = vectorizer.fit_transform([sentiment_text])
                    feature_names = vectorizer.get_feature_names_out()
                    tfidf_scores = tfidf_matrix.toarray()[0]
                    sentiment_keywords = sorted(zip(feature_names, tfidf_scores), 
                                              key=lambda x: x[1], reverse=True)
                    
                    sentiment_kr = {'positive': '긍정', 'negative': '부정', 'neutral': '중립'}[sentiment]
                    keywords_str = ', '.join([word for word, score in sentiment_keywords[:8]])
                    all_data.append([f"{sentiment_kr} 키워드", keywords_str, f"{sentiment_kr} 리뷰의 특징적 키워드"])
                except:
                    sentiment_kr = {'positive': '긍정', 'negative': '부정', 'neutral': '중립'}[sentiment]
                    all_data.append([f"{sentiment_kr} 키워드", '키워드 추출 실패', ''])
    
    all_data.append(['', '', ''])
    
    # 8. 상세 통계
    all_data.append(['[상세 통계]', '', ''])
    
    # 텍스트 길이 통계
    length_stats = analyzer.df['cleaned_review'].str.len().describe()
    all_data.append(['텍스트 길이 평균', f"{length_stats['mean']:.1f}자", ''])
    all_data.append(['텍스트 길이 표준편차', f"{length_stats['std']:.1f}자", ''])
    all_data.append(['텍스트 길이 최솟값', f"{length_stats['min']:.0f}자", ''])
    all_data.append(['텍스트 길이 중간값', f"{length_stats['50%']:.0f}자", ''])
    all_data.append(['텍스트 길이 최댓값', f"{length_stats['max']:.0f}자", ''])
    
    # 키워드 수 통계
    token_counts = analyzer.df['tokens'].apply(len)
    token_stats = token_counts.describe()
    all_data.append(['키워드 수 평균', f"{token_stats['mean']:.1f}개", ''])
    all_data.append(['키워드 수 중간값', f"{token_stats['50%']:.0f}개", ''])
    all_data.append(['키워드 수 최댓값', f"{token_stats['max']:.0f}개", ''])
    
    all_data.append(['', '', ''])
    
    # 9. 샘플 리뷰 데이터 (상위 10개)
    all_data.append(['[샘플 리뷰 데이터 (상위 10개)]', '', ''])
    all_data.append(['원본 리뷰', '규칙기반 감성', '추출된 키워드'])
    
    sample_df = analyzer.df.head(10)
    for idx, row in sample_df.iterrows():
        original_review = str(row[text_col])[:100] + "..." if len(str(row[text_col])) > 100 else str(row[text_col])
        sentiment = row['sentiment_rule']
        keywords = row['tokens_str'][:50] + "..." if len(row['tokens_str']) > 50 else row['tokens_str']
        all_data.append([original_review, sentiment, keywords])
    
    # 데이터프레임 생성 및 엑셀 저장
    df_report = pd.DataFrame(all_data, columns=['항목', '값', '설명/추가정보'])
    
    # ExcelWriter로 저장
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_report.to_excel(writer, sheet_name='쿠팡_리뷰_분석_리포트', index=False)
        
        # 워크시트 스타일링
        worksheet = writer.sheets['쿠팡_리뷰_분석_리포트']
        
        # 컬럼 폭 조정
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 40
        worksheet.column_dimensions['C'].width = 60
        
        # 헤더 스타일링
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, 4):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 섹션 헤더 스타일링 (대괄호로 시작하는 행들)
        section_font = Font(bold=True, size=11, color="0066CC")
        section_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        for row in range(2, len(all_data) + 2):
            cell_value = worksheet.cell(row=row, column=1).value
            if cell_value and str(cell_value).startswith('[') and str(cell_value).endswith(']'):
                for col in range(1, 4):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = section_font
                    cell.fill = section_fill
    
    print(f"✅ 분석 결과가 하나의 엑셀 시트로 저장되었습니다:")
    print(f"📁 파일 경로: {output_file}")
    
    return output_file

def create_summary_sheet(writer, analyzer, initial_count, final_count, text_col, rating_col):
    """요약 정보 시트 생성"""
    
    summary_data = []
    
    # 기본 정보
    summary_data.append(['분석 항목', '값', '설명'])
    summary_data.append(['분석 일시', datetime.now().strftime("%Y-%m-%d %H:%M:%S"), '분석을 수행한 날짜와 시간'])
    summary_data.append(['총 리뷰 수 (원본)', f"{initial_count:,}개", '처리 전 전체 리뷰 개수'])
    summary_data.append(['분석된 리뷰 수', f"{final_count:,}개", '전처리 후 실제 분석된 리뷰 개수'])
    summary_data.append(['제거된 리뷰 수', f"{initial_count - final_count:,}개", '빈 내용 등으로 제거된 리뷰'])
    summary_data.append(['리뷰 텍스트 컬럼', text_col, '분석에 사용된 리뷰 텍스트 컬럼명'])
    summary_data.append(['평점 컬럼', rating_col if rating_col else '없음', '분석에 사용된 평점 컬럼명'])
    
    # 텍스트 통계
    avg_length = analyzer.df['cleaned_review'].str.len().mean()
    max_length = analyzer.df['cleaned_review'].str.len().max()
    min_length = analyzer.df['cleaned_review'].str.len().min()
    
    summary_data.append(['평균 리뷰 길이', f"{avg_length:.1f}자", '전처리된 리뷰의 평균 글자 수'])
    summary_data.append(['최대 리뷰 길이', f"{max_length}자", '가장 긴 리뷰의 글자 수'])
    summary_data.append(['최소 리뷰 길이', f"{min_length}자", '가장 짧은 리뷰의 글자 수'])
    
    # 평점 통계
    if rating_col and rating_col in analyzer.df.columns:
        avg_rating = analyzer.df[rating_col].mean()
        summary_data.append(['평균 평점', f"{avg_rating:.2f}", '전체 리뷰의 평균 평점'])
    
    # 감성 분포
    sentiment_counts = analyzer.df['sentiment_rule'].value_counts()
    summary_data.append(['긍정 리뷰 수', f"{sentiment_counts.get('positive', 0):,}개", '규칙 기반 감성분석 결과'])
    summary_data.append(['부정 리뷰 수', f"{sentiment_counts.get('negative', 0):,}개", '규칙 기반 감성분석 결과'])
    summary_data.append(['중립 리뷰 수', f"{sentiment_counts.get('neutral', 0):,}개", '규칙 기반 감성분석 결과'])
    
    summary_df = pd.DataFrame(summary_data[1:], columns=summary_data[0])
    summary_df.to_excel(writer, sheet_name='1_분석요약', index=False)

def create_sentiment_analysis_sheet(writer, analyzer):
    """감성 분석 결과 시트 생성"""
    
    # 규칙 기반 감성 분석 결과
    sentiment_rule_counts = analyzer.df['sentiment_rule'].value_counts()
    sentiment_rule_ratio = (sentiment_rule_counts / len(analyzer.df) * 100).round(1)
    
    sentiment_data = []
    sentiment_data.append(['감성', '개수', '비율(%)', '분석방법'])
    
    for sentiment in ['positive', 'negative', 'neutral']:
        count = sentiment_rule_counts.get(sentiment, 0)
        ratio = sentiment_rule_ratio.get(sentiment, 0)
        sentiment_data.append([sentiment, count, ratio, '규칙기반'])
    
    # 평점 기반 감성 분석 결과 (있는 경우)
    if 'sentiment_rating' in analyzer.df.columns:
        sentiment_rating_counts = analyzer.df['sentiment_rating'].value_counts()
        sentiment_rating_ratio = (sentiment_rating_counts / len(analyzer.df) * 100).round(1)
        
        for sentiment in ['positive', 'negative', 'neutral']:
            count = sentiment_rating_counts.get(sentiment, 0)
            ratio = sentiment_rating_ratio.get(sentiment, 0)
            sentiment_data.append([sentiment, count, ratio, '평점기반'])
    
    sentiment_df = pd.DataFrame(sentiment_data[1:], columns=sentiment_data[0])
    sentiment_df.to_excel(writer, sheet_name='2_감성분석결과', index=False)

def create_keyword_analysis_sheet(writer, keywords_tfidf, keywords_krwordrank):
    """키워드 분석 시트 생성"""
    
    # TF-IDF 키워드
    tfidf_data = []
    tfidf_data.append(['순위', '키워드', 'TF-IDF 점수'])
    
    for i, (word, score) in enumerate(keywords_tfidf[:30], 1):
        tfidf_data.append([i, word, round(score, 4)])
    
    tfidf_df = pd.DataFrame(tfidf_data[1:], columns=tfidf_data[0])
    
    # KR-WordRank 키워드 (있는 경우)
    if keywords_krwordrank:
        krwordrank_data = []
        krwordrank_data.append(['순위', '키워드', 'WordRank 점수'])
        
        for i, (word, score) in enumerate(keywords_krwordrank[:30], 1):
            krwordrank_data.append([i, word, round(score, 1)])
        
        krwordrank_df = pd.DataFrame(krwordrank_data[1:], columns=krwordrank_data[0])
        
        # 두 결과를 나란히 배치
        combined_df = pd.concat([tfidf_df, pd.DataFrame([''] * len(tfidf_df)), krwordrank_df], axis=1)
        combined_df.columns = ['TF-IDF 순위', 'TF-IDF 키워드', 'TF-IDF 점수', '', 
                              'WordRank 순위', 'WordRank 키워드', 'WordRank 점수']
    else:
        combined_df = tfidf_df
    
    combined_df.to_excel(writer, sheet_name='3_키워드분석', index=False)

def create_topic_modeling_sheet(writer, topics):
    """토픽 모델링 시트 생성"""
    
    topic_data = []
    topic_data.append(['토픽 번호', '주요 키워드', '키워드 설명'])
    
    for i, topic_words in enumerate(topics, 1):
        keywords = ', '.join(topic_words[:10])
        description = f"토픽 {i}의 주요 키워드들"
        topic_data.append([f"토픽 {i}", keywords, description])
    
    topic_df = pd.DataFrame(topic_data[1:], columns=topic_data[0])
    topic_df.to_excel(writer, sheet_name='4_토픽모델링', index=False)

def create_sentiment_keywords_sheet(writer, analyzer):
    """감성별 키워드 시트 생성"""
    
    sentiment_keywords_data = []
    sentiment_keywords_data.append(['감성', '순위', '키워드', 'TF-IDF 점수'])
    
    for sentiment in ['positive', 'negative', 'neutral']:
        if sentiment in analyzer.df['sentiment_rule'].values:
            sentiment_data = analyzer.df[analyzer.df['sentiment_rule'] == sentiment]
            if len(sentiment_data) > 0:
                sentiment_text = ' '.join(sentiment_data['tokens_str'])
                vectorizer = TfidfVectorizer(max_features=20, token_pattern=r'\b\w+\b')
                try:
                    tfidf_matrix = vectorizer.fit_transform([sentiment_text])
                    feature_names = vectorizer.get_feature_names_out()
                    tfidf_scores = tfidf_matrix.toarray()[0]
                    sentiment_keywords = sorted(zip(feature_names, tfidf_scores), 
                                              key=lambda x: x[1], reverse=True)
                    
                    for i, (word, score) in enumerate(sentiment_keywords, 1):
                        sentiment_keywords_data.append([sentiment, i, word, round(score, 4)])
                except:
                    sentiment_keywords_data.append([sentiment, 1, '키워드 추출 실패', 0])
    
    sentiment_keywords_df = pd.DataFrame(sentiment_keywords_data[1:], columns=sentiment_keywords_data[0])
    sentiment_keywords_df.to_excel(writer, sheet_name='5_감성별키워드', index=False)

def create_detailed_reviews_sheet(writer, analyzer):
    """상세 리뷰 데이터 시트 생성"""
    
    # 필요한 컬럼만 선택
    columns_to_include = [analyzer.text_col, 'cleaned_review', 'sentiment_rule']
    
    if analyzer.rating_col and analyzer.rating_col in analyzer.df.columns:
        columns_to_include.append(analyzer.rating_col)
        columns_to_include.append('sentiment_rating')
    
    columns_to_include.extend(['tokens_str'])
    
    detailed_df = analyzer.df[columns_to_include].copy()
    
    # 컬럼명 변경
    column_names = {
        analyzer.text_col: '원본_리뷰',
        'cleaned_review': '전처리된_리뷰',
        'sentiment_rule': '규칙기반_감성',
        'tokens_str': '추출된_키워드'
    }
    
    if analyzer.rating_col and analyzer.rating_col in analyzer.df.columns:
        column_names[analyzer.rating_col] = '평점'
        column_names['sentiment_rating'] = '평점기반_감성'
    
    detailed_df = detailed_df.rename(columns=column_names)
    
    # 엑셀 시트 크기 제한을 고려하여 최대 10000개까지만
    if len(detailed_df) > 10000:
        detailed_df = detailed_df.head(10000)
    
    detailed_df.to_excel(writer, sheet_name='6_상세리뷰데이터', index=False)

def create_statistics_sheet(writer, analyzer, rating_col):
    """통계 요약 시트 생성"""
    
    stats_data = []
    stats_data.append(['구분', '항목', '값'])
    
    # 텍스트 길이 통계
    length_stats = analyzer.df['cleaned_review'].str.len().describe()
    stats_data.append(['텍스트 길이', '평균', f"{length_stats['mean']:.1f}자"])
    stats_data.append(['텍스트 길이', '표준편차', f"{length_stats['std']:.1f}자"])
    stats_data.append(['텍스트 길이', '최솟값', f"{length_stats['min']:.0f}자"])
    stats_data.append(['텍스트 길이', '25% 분위수', f"{length_stats['25%']:.0f}자"])
    stats_data.append(['텍스트 길이', '중간값', f"{length_stats['50%']:.0f}자"])
    stats_data.append(['텍스트 길이', '75% 분위수', f"{length_stats['75%']:.0f}자"])
    stats_data.append(['텍스트 길이', '최댓값', f"{length_stats['max']:.0f}자"])
    
    # 평점 통계 (있는 경우)
    if rating_col and rating_col in analyzer.df.columns:
        rating_stats = analyzer.df[rating_col].describe()
        stats_data.append(['평점', '평균', f"{rating_stats['mean']:.2f}"])
        stats_data.append(['평점', '표준편차', f"{rating_stats['std']:.2f}"])
        stats_data.append(['평점', '최솟값', f"{rating_stats['min']:.0f}"])
        stats_data.append(['평점', '중간값', f"{rating_stats['50%']:.0f}"])
        stats_data.append(['평점', '최댓값', f"{rating_stats['max']:.0f}"])
        
        # 평점별 분포
        rating_counts = analyzer.df[rating_col].value_counts().sort_index()
        for rating, count in rating_counts.items():
            ratio = count / len(analyzer.df) * 100
            stats_data.append(['평점 분포', f"{rating}점", f"{count}개 ({ratio:.1f}%)"])
    
    # 토큰 수 통계
    token_counts = analyzer.df['tokens'].apply(len)
    token_stats = token_counts.describe()
    stats_data.append(['키워드 수', '평균', f"{token_stats['mean']:.1f}개"])
    stats_data.append(['키워드 수', '중간값', f"{token_stats['50%']:.0f}개"])
    stats_data.append(['키워드 수', '최댓값', f"{token_stats['max']:.0f}개"])
    
    stats_df = pd.DataFrame(stats_data[1:], columns=stats_data[0])
    stats_df.to_excel(writer, sheet_name='7_통계요약', index=False)

# =============================================================================
# 메인 실행 함수
# =============================================================================

def main():
    """메인 실행 함수"""
    
    print("🚀 쿠팡 리뷰 분석 시스템 시작")
    print("=" * 50)
    
    # 1. 데이터 로딩
    df = analyze_coupang_reviews()
    if df is None:
        return
    
    print(f"✅ 데이터 로딩 성공: {len(df):,}개 리뷰")
    print(f"📊 컬럼 정보: {list(df.columns)}")
    
    # 2. 컬럼 분석
    text_columns, rating_columns = identify_columns(df)
    
    print(f"🔍 텍스트 컬럼 후보: {text_columns}")
    print(f"🔍 평점 컬럼 후보: {rating_columns}")
    
    # 3. 컬럼 자동 선택
    if text_columns:
        selected_text_col = text_columns[0]  # 첫 번째 텍스트 컬럼 선택
        print(f"📝 선택된 리뷰 텍스트 컬럼: '{selected_text_col}'")
    else:
        print("❌ 리뷰 텍스트 컬럼을 찾을 수 없습니다.")
        return
    
    if rating_columns:
        selected_rating_col = rating_columns[0]  # 첫 번째 평점 컬럼 선택
        print(f"⭐ 선택된 평점 컬럼: '{selected_rating_col}'")
    else:
        selected_rating_col = None
        print("⚠️ 평점 컬럼 없음 (텍스트 기반 분석만 수행)")
    
    print("\n🔄 분석 시작...")
    print("=" * 50)
    
    # 4. 분석 실행 및 엑셀 출력
    analyzer = run_analysis_with_identified_columns(df, selected_text_col, selected_rating_col)
    
    print("=" * 50)
    print("🎉 분석 완료!")
    
    return analyzer

if __name__ == "__main__":
    analyzer = main()